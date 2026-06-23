import time
import subprocess
import os
import re
import threading

from hypium.model import UiParam
from sympy import swinnerton_dyer_poly
from openpyxl import load_workbook, Workbook

from .harmony_driver import HarmonyDriver
from .utils import check_stop, UiDriver, By, MatchPattern, log_and_emit
from .config import Config

_EXCEL_LOCK = threading.Lock()

_RECORDER_BUNDLE = 'com.huawei.hmos.screenrecorder'
_RECORDER_ABILITY = 'com.huawei.hmos.screenrecorder.ServiceExtAbility'


class Xiaoyilivechat(HarmonyDriver):

    def _hdc_shell(self, device_sn, *args):
        return subprocess.run(
            ['hdc', '-t', device_sn, 'shell'] + list(args),
            check=False, capture_output=True, text=True, timeout=10
        )

    def initialize(self, device_sn, task_id=None, test_case_id=None, **kwargs) -> bool:
        if not super().initialize(device_sn, task_id=task_id, test_case_id=test_case_id, **kwargs):
            return False
        driver = self._get_driver(device_sn)
        if not driver:
            return False
        # 点开小艺聊天窗口
        user_center = driver.find_component(By.text("小艺"))
        if user_center:
            user_center.click()
            time.sleep(2)
        # 清除上下文
        # 进入设置界面
        driver.touch(By.isAfter(By.type('Image')).isBefore(By.key('water_mark.build.stack')).type('SymbolGlyph'))
        driver.wait(2)
        clear_text=driver.find_component(By.text('清除上下文'))
        if  clear_text:
            clear_text.click()
            time.sleep(2)
        # 进入设置界面删除对话记录
        driver.touch(By.isAfter(By.type('Image')).isBefore(By.key('water_mark.build.stack')).type('SymbolGlyph'))
        driver.wait(2)
        driver.swipe(UiParam.UP,30,side=UiParam.LEFT)
        clear_chat=driver.find_component(By.text("删除对话记录"))
        if  clear_chat:
            clear_chat.click()
            time.sleep(1)
            # 确认删除
            delete_button=driver.find_component(By.text("删除"))
            if delete_button:
                delete_button.click()
                time.sleep(1)
        return True

    def pre_proc   ess(self, device_sn, task_id=None, test_case_id=None, **kwargs) -> bool:
        # 开启通话聊天
        driver = self._get_driver(device_sn)
        driver.touch(By.isAfter(By.key('ChatTitleMenu')).isBefore(By.key('title_bar.broadcastType.icon')).type(
            'SymbolGlyph'))
        driver.wait(2)
        try:
            if driver.find_component(By.text("小艺")):
                self._log(level='DEBUG', content="成功进行通话", task_id=task_id, test_case_id=test_case_id)
        except Exception:
            self._log(level='ERROR', content="通话失败", task_id=task_id, test_case_id=test_case_id)
            return False
        # 开启录屏
        self._record_file_name = f"{test_case_id}.mp4"
        self._hdc_shell(device_sn, 'aa', 'start', '-b', _RECORDER_BUNDLE, '-a', _RECORDER_ABILITY,
                        '--ps', 'CustomizedFileName', self._record_file_name)
        self._log(level='INFO', content=f"启动录屏: {self._record_file_name}", task_id=task_id,
                  test_case_id=test_case_id)
        driver.wait(2)
        return True

    def post_process(self, device_sn, task_id=None, test_case_id=None, **kwargs) -> bool:
        driver = self._get_driver(device_sn)
        # 停止录屏
        self._hdc_shell(device_sn, 'aa', 'start', '-b', _RECORDER_BUNDLE, '-a', _RECORDER_ABILITY)
        self._log(level='INFO', content="停止录屏", task_id=task_id, test_case_id=test_case_id)
        time.sleep(5)
        # 通话挂断
        try:
            hangup_btn = driver.find_component(
                By.isAfter(By.key('live.tool_bar.hangup_button')).isBefore(By.key('GuideText')).type('SymbolGlyph'))
            if hangup_btn:
                hangup_btn.click()
        except Exception as e:
            self._log(level='WARNING', content=f"挂断通话失败: {e}", task_id=task_id, test_case_id=test_case_id)
        driver.wait(5)
        texts = driver.find_all_components(By.type("Text"))
        text=[t.getText() for t in texts]
        self._question_text=text[3]
        self._answer_text=text[5]
        print(text)
        print(text[3])
        print(text[5])
        # for text in texts:
        #     s = str(text)
        #     print(s.split("##")[1].split("#U")[0] if "##" in s else s.split("#")[0][3].split("#U")[0])
        return True



    def get_results(self, device_sn, task_id=None, test_case_id=None, **kwargs) -> dict:
        record_file_name = getattr(self, '_record_file_name', 'record.mp4')
        question_text=getattr(self, '_question_text', None)
        answer_text=getattr(self, '_answer_text', None)
        try:
            query = self._hdc_shell(device_sn, 'mediatool', 'query', record_file_name, '-u')
            lines = query.stdout.strip().split('\n')
            device_path = lines[1].strip() if len(lines) > 1 else ''
            if 'uri' in query.stdout and len(lines) > 2:
                subprocess.run(
                    ['hdc', '-t', device_sn, 'shell', 'mediatool', 'recv', lines[2].strip(), '/data/local/tmp'],
                    check=False, capture_output=True, text=True, timeout=120
                )
                device_path = f'/data/local/tmp/{record_file_name}'
            local_dir = os.path.join(Config.STATIC_BASE_PATH, 'case_result',
                                     task_id or 'default_task_id',
                                     test_case_id or 'default_id', device_sn)
            os.makedirs(local_dir, exist_ok=True)
            local_path = os.path.join(local_dir, record_file_name)
            subprocess.run(['hdc', '-t', device_sn, 'file', 'recv', device_path, local_path],
                           check=False, capture_output=True, text=True, timeout=120)
            self._append_to_excel(task_id, test_case_id, question_text, answer_text)
            return {'success': True, 'message': 'Success', 'record_path': local_path,'question':question_text,'answer':answer_text}
        except Exception as e:
            self._log(level='ERROR', content=f"获取录屏文件失败: {e}", task_id=task_id, test_case_id=test_case_id)
            return {'success': False, 'message': str(e), 'record_path': '','question':"",'answer':''}

    def _append_to_excel(self, task_id, test_case_id, question_text, answer_text):
        with _EXCEL_LOCK:
            try:
                excel_dir = os.path.join(Config.STATIC_BASE_PATH, 'case_result',
                                         task_id or 'default_task_id')
                os.makedirs(excel_dir, exist_ok=True)
                excel_path = os.path.join(excel_dir, 'xiaoyi_chat_results.xlsx')
                if os.path.exists(excel_path):
                    wb = load_workbook(excel_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['test_case_id', 'question_text', 'answer_text'])
                ws.append([test_case_id, question_text, answer_text])
                wb.save(excel_path)
                self._log(level='INFO', content=f"结果已写入Excel: {excel_path}",
                          task_id=task_id, test_case_id=test_case_id)
            except Exception as e:
                self._log(level='ERROR', content=f"写入Excel失败: {e}",
                          task_id=task_id, test_case_id=test_case_id)
